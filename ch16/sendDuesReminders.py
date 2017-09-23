#Mac
#! /usr/local/var/lib/pyenv/versions/anaconda3-2.0.1/python3.4
#Linux
# #! /usr/bin/python3

## sendDuesReminders.py - 16.5 スプレッドシートの支払状況に基づき、メールを送信する
# Usage:
# python sendDuesReminders.py [your Gmail password]

import openpyxl
import smtplib
import sys

# スプレッドシートを開き最近の支払状況を取得する
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

last_col = sheet.get_highest_column()
latest_month = sheet.cell(row=1, column=last_col).value

# 会員の支払状況を調べる
unpaid_members = {}
for r in range(2, sheet.get_highest_row() + 1):
	payment = sheet.cell(row=r, column=last_col).value	# 最新月（要はいちばん右の列）の値を格納
	if payment != 'paid':
		name = sheet.cell(row=r, column=1).value
		email = sheet.cell(row=r, column=2).value
		unpaid_members[name] = email

# メールアカウントにログインする
##########
#とりあえずSMTPものは不用意にメールを送信したくないので、コメントアウトする
##########
#smtp_obj = smtplib.SMTP('smtp.gmail.com', 587)
#smtp_obj.ehlo()
#smtp_obj.starttls()
#smtp_obj.login('my_email_address@gmail.com', sys.argv[1])	# argv[1]はpasswerd

# リマインダーメールを送信する
for name, email in unpaid_members.items():
	body = """Subject: {} dues unpaid.
Dear {},
Records show that you have not paid dues for {}. Please make this payment as soon as possible. Thank you!
""".format(latest_month, name, latest_month)
	print('メール送信中 {}...'.format(email))
#	sendmeil_status = smtp_obj.sendmail('my_email_address@gmail.com', email, body)

#	if sendmeil_status != {}:
#		print('{}へメール送信中に問題が起こりました：{}'.format(email, sendmeil_status))

#smtp_obj.quit()


