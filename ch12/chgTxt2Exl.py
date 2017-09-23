#Windows
# #! python3
#Mac
#! /usr/local/var/lib/pyenv/versions/anaconda3-2.0.1/python3.4
#Linux
# #! /usr/bin/python3

## chgTxt2Exl.py - 12.14.4 カレントフォルダの全テキストファイルをスプレッドシートに変換する
# Usage:
# python chgTxt2Exl.py

import os
import openpyxl
import logging

logging.basicConfig(level=logging.DEBUG,
#	filename='myProgramLog.txt',
	format=' %(asctime)s - %(levelname)s - %(message)s')

# Workbookを開く
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

n = 0
for filename in sorted(os.listdir('.')):
	# 拡張子が.txtでなければスキップ
	if not filename.lower().endswith('.txt'):
		continue

	logging.debug('filename：{}'.format(filename))

	# ファイルごとに列番号を1,2,3,...と増やしていく
	n += 1
	text_file = open(filename, 'r', encoding='utf-8')
	lines = text_file.readlines()
	text_file.close()

	# 1行目にファイル名を入れてみる
	new_sheet.cell(row = 1, column=n).value = filename
	row = 2
	# 行をコピーしていく
	for line in lines:
		new_sheet.cell(row = row, column = n).value = line
		row += 1

# Excelファイルを保存する
new_wb.save('texts.xlsx')
