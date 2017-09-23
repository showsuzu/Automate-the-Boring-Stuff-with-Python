#Windows
# #! python3
#Mac
#! /usr/local/var/lib/pyenv/versions/anaconda3-2.0.1/python3.4
#Linux
# #! /usr/bin/python3

## chgExl2Txt.py - 12.14.5 スプレッドシートからテキストファイルに変換する
# Usage:
# python chgExl2Txt.py <excel filename>

import sys
import openpyxl
import logging

logging.basicConfig(level=logging.DEBUG,
#	filename='myProgramLog.txt',
	format=' %(asctime)s - %(levelname)s - %(message)s')

# 引数をチェックする
logging.debug('入力数：{}'.format(len(sys.argv)))
if len(sys.argv) != 2:
	print('使い方: python chgExl2Txt.py <excel filename>')
	sys.exit()

filename = sys.argv[1]

# Workbookを開く
wb = openpyxl.load_workbook(filename)
sheet = wb.active

for col in range(1, sheet.max_column + 1):
	# 元ファイル名に列番号をつけてテキストファイル名とする
	txtname = '{}_{:03}.txt'.format(filename, col)
	text_file = open(txtname, 'w', encoding='utf-8')
	for row in range(1, sheet.max_row + 1):
		value = sheet.cell(column = col, row = row).value
		if value != None:
			text_file.write(str(value))
			text_file.write('\n')
	text_file.close()
