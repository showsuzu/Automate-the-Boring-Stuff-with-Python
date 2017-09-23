#Windows
# #! python3
#Mac
#! /usr/local/var/lib/pyenv/versions/anaconda3-2.0.1/python3.4
#Linux
# #! /usr/bin/python3

## blankRowInserter.py - 12.14.2 指定したExcelに、指定した行から指定した行の空行を挿入する
# Usage:
# python blankRowInserter.py <N> <M> <filename>

import sys
import openpyxl
import logging

logging.basicConfig(level=logging.CRITICAL,
#	filename='myProgramLog.txt',
	format=' %(asctime)s - %(levelname)s - %(message)s')

# 引数をチェックする
if len(sys.argv) != 4:
	print('使い方: python blankRowInserter.py  <空行開始行> <空行数> <filename>')
	sys.exit()

n = int(sys.argv[1])
m = int(sys.argv[2])
filename = sys.argv[3]

# Workbookを開く
wb = openpyxl.load_workbook(filename)
sheet = wb.active

new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# sheetの最大行まで操作する
for row in range(1, sheet.max_row + 1):
	# 現在行が空行の開始か否か
	if row < n:
		new_row = row
	else:
		new_row = row + m  # m行分の空行を入れる
	# 全ての列をコピーする
	for col in range(1, sheet.max_column + 1):
		old_cell = sheet.cell(row = row, column = col)
		new_cell = new_sheet.cell(row = new_row, column = col)
		new_cell.value = old_cell.value	# 既存ファイルのセルを、新シートのセルにコピー

new_wb.save(filename + '.ins.xlsx')
