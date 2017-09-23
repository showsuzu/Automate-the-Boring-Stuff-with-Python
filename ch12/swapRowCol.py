#Windows
# #! python3
#Mac
#! /usr/local/var/lib/pyenv/versions/anaconda3-2.0.1/python3.4
#Linux
# #! /usr/bin/python3

## swapRowCol.py - 12.14.3 指定したExcelの行と列を入れ替える
# Usage:
# python swapRowCol.py <filename>

import sys
import openpyxl
import logging

logging.basicConfig(level=logging.CRITICAL,
#	filename='myProgramLog.txt',
	format=' %(asctime)s - %(levelname)s - %(message)s')

# 引数をチェックする
logging.debug('入力数：{}'.format(len(sys.argv)))
if len(sys.argv) != 2:
	print('使い方: python swapRowCol.py <filename>')
	sys.exit()

filename = sys.argv[1]

# Workbookを開く
wb = openpyxl.load_workbook(filename)
sheet = wb.active

new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# sheetの最大行まで操作する
for row in range(1, sheet.max_row + 1):
	# 行の列を操作する
	for col in range(1, sheet.max_column + 1):
		# 値を入れ替える
		val = sheet.cell(row = row, column = col).value
		new_sheet.cell(row = col, column = row).value = val

new_wb.save(filename + '.swp.xlsx')
