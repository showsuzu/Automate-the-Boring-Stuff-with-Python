#Windows
# #! python3
#Mac
#! /usr/local/var/lib/pyenv/versions/anaconda3-2.0.1/python3.4
#Linux
# #! /usr/bin/python3

## updateProduce.py - 12.6 農産物スプレッドシートの価格を訂正する
# Usage:
# python updateProduce.py

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# 農産物の種類と、更新する価格
PRICE_UPDATES = {	'Garlic': 3.07,
					'Celery': 1.19,
					'Lemon' : 1.27}

# 行をループして価格を更新する
for row_num in range(2, sheet.max_row):		# 先頭行はタイトルなのでスキップする
	produce_name = sheet.cell(row = row_num, column = 1).value
	if produce_name in PRICE_UPDATES:		# 更新する価格用の辞書のキーに、produce_nameが存在するか？
		sheet.cell(row = row_num, column = 2).value = PRICE_UPDATES[produce_name]	# 元のSheetのB列の価格を更新する

wb.save('updatedProduceSales.xlsx')
