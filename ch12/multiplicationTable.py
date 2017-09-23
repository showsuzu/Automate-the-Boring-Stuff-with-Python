#Windows
# #! python3
#Mac
#! /usr/local/var/lib/pyenv/versions/anaconda3-2.0.1/python3.4
#Linux
# #! /usr/bin/python3

## multiplicationTable.py - 12.14.1 掛け算の表を作成する
# Usage:
# python multiplicationTable.py [N]

import sys
import openpyxl
from openpyxl.styles import Font
import logging

logging.basicConfig(level=logging.CRITICAL,
#	filename='myProgramLog.txt',
	format=' %(asctime)s - %(levelname)s - %(message)s')

# 引数をチェックする
if len(sys.argv) == 2:
	nums = int(sys.argv[1])
	if nums < 1:
		print('入力する値は、１よりも大きい値にして下さい')
		sys.exit()
else:
	print('使い方: python multiplicationTable.py <表の最大数>')
	sys.exit()

# Workbookを開く
wb = openpyxl.Workbook()
sheet = wb.active

# ウィンドウ枠を固定する
sheet.freeze_panes = 'B2'

# 太字フォント
font1 = Font(bold = True)

# A2から縦に数値を１から並べ、B1から横に数値を並べる
# （入力値は１よりも大きい前提）
for num in range(1, nums + 1):
	pos = num + 1
	# 行の見出しの数値と書式を設定
	sheet.cell(row=pos, column=1).value = num	# A2から縦に入力値を並べる
	sheet.cell(row=pos, column=1).font = font1	# 太字指定
	# 列の見出しの数値と書式を設定
	sheet.cell(row=1, column=pos).value = num	# B1から横に入力値を並べる
	sheet.cell(row=1, column=pos).font = font1	# 太字指定

# 掛け算の代入
for i in range(1, nums + 1):					# row
	for j in range(1, nums + 1):				# column
		# 掛け算の代入
		cell = sheet.cell(row = i + 1, column = j + 1)
#		# 計算して値を代入する
#		cell.value = i * j

		# 計算式を代入する
		val1_cell = sheet.cell(row = i + 1, column = 1)
		val2_cell = sheet.cell(row = 1, column = j + 1)
		calc = '=' + val1_cell.coordinate + ' * ' + val2_cell.coordinate	# 計算式を作成する
		logging.debug('CELL：{}　計算式：{}'.format(cell.coordinate, calc))
		cell.value = calc	# 計算式を代入

# Workbookを保存する
wb.save('NxNtable.xlsx')
