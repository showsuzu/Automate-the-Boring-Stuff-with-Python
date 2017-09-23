#Windows
# #! python3
#Mac
#! /usr/local/var/lib/pyenv/versions/anaconda3-2.0.1/python3.4
#Linux
# #! /usr/bin/python3

## excel2csv.py - 14.8.1 カレントフォルダのExcelをCSVへ変換する
# Usage:
# python excel2csv.py

import openpyxl
import sys
import os
import csv

# カレントディレクトリの全ファイルをループする
for xls_filename in os.listdir('.'):
	if not xls_filename.endswith('.xlsx'):
		continue	# CSVファイルでなければスキップする

	wb = openpyxl.load_workbook(xls_filename)
	for sheet_name in wb.get_sheet_names():
		# sheet_nameからsheet内容を取得する
		sheet = wb.get_sheet_by_name(sheet_name)

		# CSVファイルを準備する
		csv_filename = xls_filename[:-5] + '_' + sheet_name + '.csv'
		csv_file = open(csv_filename, 'w', encoding='utf-8', newline='')
		csv_writer = csv.writer(csv_file)

		for row_num in range(1, sheet.max_row + 1):
			row_data = [] # セルをこのリストに追加する
			# 行のセルをループする
			for col_num in range(1, sheet.max_column + 1):
				# セルをrow_dataに追加する
				row_data.append(sheet.cell(column=col_num, row=row_num).value)
			# row_dataをCSVファイルに書き出す
			csv_writer.writerow(row_data)

		csv_file.close()
